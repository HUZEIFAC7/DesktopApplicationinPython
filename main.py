import sys
import pandas as pd
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QProgressBar, QMessageBox
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

class ExcelProcessor(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.excel_data = None
        self.processed_excel_data = {}

    def initUI(self):
        self.setWindowTitle('Excel Processor')
        self.setGeometry(100, 100, 400, 200)
        layout = QVBoxLayout(self)

        self.upload_btn = QPushButton('Upload Excel File', self)
        self.upload_btn.clicked.connect(self.upload_file)
        layout.addWidget(self.upload_btn)

        self.download_btn = QPushButton('Download Processed File', self)
        self.download_btn.clicked.connect(self.download_file)
        self.download_btn.setEnabled(False)
        layout.addWidget(self.download_btn)

        self.progress = QProgressBar(self)
        layout.addWidget(self.progress)

    def upload_file(self):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, 'Open Excel File', '', 'Excel Files (*.xlsx *.xls)', options=options)
        if file_path:
            self.load_data(file_path)

    def load_data(self, file_path):
        self.progress.setValue(0)
        try:
            # Load Excel file into a dictionary of dataframes (one for each sheet)
            self.excel_data = pd.read_excel(file_path, sheet_name=None, header=None, engine='openpyxl')
            self.process_data()
            self.progress.setValue(100)
            self.download_btn.setEnabled(True)
            QMessageBox.information(self, 'Success', 'File processed successfully!')
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to load file:\n{str(e)}')

    def process_data(self):
        month_order = {month: i for i, month in enumerate(["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])}

        monthly_sheets = {}

        for sheet_name, df in self.excel_data.items():
            # Fix headers (third row as the header)
            df.columns = df.iloc[2]  # Set the third row as headers
            df = df.drop([0, 1, 2])  # Drop the first three rows
            df = df.reset_index(drop=True)

            # Rename columns to correct structure
            df.columns = ['Date', 'Type', 'Particulars', 'Bank', 'Instrument no.', 'Instrument date', 'Status', 'Amount']

            # Ensure date columns are parsed correctly
            df['Instrument date'] = pd.to_datetime(df['Instrument date'], errors='coerce')
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce')

            # Group by year and month from 'Instrument date'
            for group, dataframe in df.groupby([df['Instrument date'].dt.year, df['Instrument date'].dt.month]):
                year, month = group
                month_name = pd.Timestamp(year=year, month=month, day=1).strftime('%b')

                # Sort by 'Type' and 'Instrument date'
                dataframe = dataframe.sort_values(by=['Type', 'Instrument date'], ascending=[False, True])

                # Format the date columns
                dataframe['Instrument date'] = dataframe['Instrument date'].dt.strftime('%d-%b-%y')
                dataframe['Date'] = dataframe['Date'].dt.strftime('%d-%b-%y')

                # Summary calculations
                summary_data = {
                    'Particulars': ['Total Cheques Received DR', 'Total Cheques Paid CR', 'Net Balance Receivable / (Payable)'],
                    'Amount': [
                        dataframe.loc[dataframe['Type'] == 'Received', 'Amount'].sum(),
                        dataframe.loc[dataframe['Type'] == 'Issued', 'Amount'].sum(),
                        dataframe.loc[dataframe['Type'] == 'Received', 'Amount'].sum() - dataframe.loc[dataframe['Type'] == 'Issued', 'Amount'].sum()
                    ],
                    'No. cheques': [
                        dataframe[dataframe['Type'] == 'Received'].shape[0],
                        dataframe[dataframe['Type'] == 'Issued'].shape[0],
                        dataframe[dataframe['Type'] == 'Received'].shape[0] - dataframe[dataframe['Type'] == 'Issued'].shape[0]
                    ]
                }
                summary_df = pd.DataFrame(summary_data)

                # Combine data with the summary
                monthly_sheets[month_name] = pd.concat([dataframe, summary_df], ignore_index=True)

        # Sort the monthly sheets by month order
        sorted_months = sorted(monthly_sheets.keys(), key=lambda x: month_order[x])

        # Create a summary sheet with totals for each month
        summary_sheet_data = {
            'Month': [],
            'Total Cheques Received DR': [],
            'Total Cheques Paid CR': [],
            'Net Balance Receivable / (Payable)': [],
            'Total Received Cheques': [],
            'Total Issued Cheques': [],
            'Net Cheques': []
        }

        for month in sorted_months:
            df = monthly_sheets[month]
            summary_sheet_data['Month'].append(month)
            summary_sheet_data['Total Cheques Received DR'].append(df.iloc[-3]['Amount'])
            summary_sheet_data['Total Cheques Paid CR'].append(df.iloc[-2]['Amount'])
            summary_sheet_data['Net Balance Receivable / (Payable)'].append(df.iloc[-1]['Amount'])
            summary_sheet_data['Total Received Cheques'].append(df.iloc[-3]['No. cheques'])
            summary_sheet_data['Total Issued Cheques'].append(df.iloc[-2]['No. cheques'])
            summary_sheet_data['Net Cheques'].append(df.iloc[-1]['No. cheques'])

        summary_sheet_df = pd.DataFrame(summary_sheet_data)

        # Add a total row
        total_row = {
            'Month': 'Total',
            'Total Cheques Received DR': summary_sheet_df['Total Cheques Received DR'].sum(),
            'Total Cheques Paid CR': summary_sheet_df['Total Cheques Paid CR'].sum(),
            'Net Balance Receivable / (Payable)': summary_sheet_df['Net Balance Receivable / (Payable)'].sum(),
            'Total Received Cheques': summary_sheet_df['Total Received Cheques'].sum(),
            'Total Issued Cheques': summary_sheet_df['Total Issued Cheques'].sum(),
            'Net Cheques': summary_sheet_df['Net Cheques'].sum()
        }
        summary_sheet_df = pd.concat([summary_sheet_df, pd.DataFrame([total_row])], ignore_index=True)

        # Store processed data for saving
        self.processed_excel_data = {
            **{month: monthly_sheets[month] for month in sorted_months},
            'Summary': summary_sheet_df
        }

    def download_file(self):
        options = QFileDialog.Options()
        save_path, _ = QFileDialog.getSaveFileName(self, 'Save Updated Excel File', '', 'Excel Files (*.xlsx *.xls)', options=options)
        if save_path:
            try:
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    for sheet_name, df in self.processed_excel_data.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Load workbook to apply formatting
                workbook_user = load_workbook(save_path)

                # Define border style
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                # Apply borders to all sheets
                for sheet_name in workbook_user.sheetnames:
                    worksheet = workbook_user[sheet_name]
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.border = thin_border

                # Save formatted workbook
                workbook_user.save(save_path)

                QMessageBox.information(self, 'Success', 'File saved and formatted successfully!')
                self.download_btn.setEnabled(False)
                self.progress.setValue(0)
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Failed to save Excel file:\n{str(e)}')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ExcelProcessor()
    ex.show()
    sys.exit(app.exec_())
